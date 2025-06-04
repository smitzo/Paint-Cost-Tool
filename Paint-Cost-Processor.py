import openpyxl
from openpyxl.styles import Font, PatternFill
import customtkinter as ctk
from tkinter import filedialog, messagebox, StringVar, Toplevel, Label, PhotoImage,Text,simpledialog
import os
import json
from PIL import Image, ImageTk
#import threading
import tkinter as tk
import random
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import hashlib


about_text=("Pdp Processor v1.0\n"
                               "Developed by Jiren Pandya\n"
                               "Other Contributor:\n"
                               "Applied by:\n"
                               "This software is the property of.\n"
                               "This software helps in analyzing Excel sheets.\n"
                               "In upcoming updates, we will provide time graph analysis, "
                               "bug fixes, and additional features.\n"
                               "The user ID is 0 and the password is 0.\n"
                               "The password and user ID are unchangeable.\n"
                               "For any help regarding this software, please call +91 8401751355 "
                               "(Jiren Pandya) or email at pandyajiren15@gmail.com.")


# Function to save settings
def save_settings():
    settings = {
        "theme_color": theme_color,
        "appearance_mode": ctk.get_appearance_mode()
    }
    with open("settings.json", "w") as f:
        json.dump(settings, f)

# Function to load settings
def load_settings():
    global theme_color
    try:
        with open("settings.json", "r") as f:
            settings = json.load(f)
            theme_color = settings.get("theme_color", "purple")
            ctk.set_appearance_mode(settings.get("appearance_mode", "Light"))
    except FileNotFoundError:
        theme_color = "purple"
        ctk.set_appearance_mode("Light") 

def upload_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        file_label.configure(text=os.path.basename(file_path), text_color=theme_color)
        global uploaded_file
        uploaded_file = file_path
        result_box.insert(ctk.END, f"File uploaded: {os.path.basename(file_path)}\n")


def save_styled_df_to_excel_with_width(file_name, styled_df):
    # Save the styled DataFrame to Excel (without the 'Color Code' column)
    styled_df.to_excel(file_name, engine='openpyxl', index=False)

    # Optionally adjust column widths if necessary
    wb = load_workbook(file_name)
    sheet = wb.active
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width
    wb.save(file_name)


def highlight_cells():
    global uploaded_file
    global styled_df  # Make styled_df global to use it in the download_file function
    global df_cleaned  # Make df_cleaned global for accessing in download_file function

    # Use the default values for cost and quantity columns and sheet name
    cost_column = cost_column_var.get()  # Dynamically get the cost column
    #print("costcolumn=" + cost_column)
    qty_column = qty_column_var.get()  # Dynamically get the quantity column
    Sheet_name = Sheet_var.get()  # Dynamically get the sheet name

    if uploaded_file:
        # Read the Excel file, skipping the first two rows
        df = pd.read_excel(uploaded_file, skiprows=2)
        titles=list(df.columns)
        #print(titles)
        titles[1],titles[2]=titles[2],titles[1]
        df=df[titles]

        # Rename columns for easier access (keep these consistent with your data)
        #df.columns = ['Row Labels', 'Average of Output Unit Cost', 'Sum of Output Qty']

        # Load workbook to check for bold text
        wb = load_workbook(uploaded_file)
        sheet = wb[Sheet_name]

        # Initialize a variable to keep track of the current main product
        current_main_product = None
        main_product_column = []

        # Iterate over the DataFrame to check for bold values
        for idx, row in df.iterrows():
            if sheet[f'A{idx + 3}'].font.bold:
                df.at[idx, 'Row Labels'] = '-'
            if sheet[f'{cost_column}{idx + 3}'].font.bold:  # Use dynamic cost column
                df.at[idx, 'Average of Output Unit Cost'] = '-'
            if sheet[f'{qty_column}{idx + 3}'].font.bold:  # Use dynamic quantity column
                df.at[idx, 'Sum of Output Qty'] = '-'

        # Filter out rows with placeholder '-'
        df_cleaned = df[df['Row Labels'] != '-'].copy()

        # Map sub-products to main products based on 'F0' logic
        for idx, row in df_cleaned.iterrows():
            if str(row['Row Labels']).startswith(('F0', 'N0')):
                current_main_product = row['Row Labels']
            main_product_column.append(current_main_product)

        # Add the 'Main Product' column to the DataFrame
        df_cleaned['Main Product'] = main_product_column

        # Check number of sub-products for each main product
        product_counts = df_cleaned['Main Product'].value_counts().to_dict()

        # For each product group, calculate the dynamic thresholds
        def calculate_thresholds(group):
            cost_min = group['Average of Output Unit Cost'].min()
            cost_max = group['Average of Output Unit Cost'].max()
            cost_mean = group['Average of Output Unit Cost'].mean()
            cost_75th = group['Average of Output Unit Cost'].quantile(0.75)
            qty_max = group['Sum of Output Qty'].max()

            return cost_min, cost_max, cost_mean, cost_75th, qty_max

        # Calculate thresholds for each product group
        thresholds = df_cleaned.groupby('Main Product', group_keys=False).apply(calculate_thresholds).to_dict()

        # Dictionary to track if a red cell has already been assigned for each product
        red_assigned = {}

        # Function to determine the color code based on dynamic thresholds and the gap rule
        def determine_color_code(row):
            product = row['Main Product']
            cost = row['Average of Output Unit Cost']
            qty = row['Sum of Output Qty']

            # Fetch thresholds for this product
            cost_min, cost_max, cost_mean, cost_75th, qty_max = thresholds.get(product, (0, 0, 0, 0, 0))

            # Initialize the tracking dictionary if not already done
            if product not in red_assigned:
                red_assigned[product] = False  # No red cell assigned yet

            # Only one sub-product - automatically green
            if product_counts[product] == 1:
                return 'green'

            # Sort costs for this main product to check differences
            product_group = df_cleaned[df_cleaned['Main Product'] == product]
            sorted_costs = sorted(product_group['Average of Output Unit Cost'].unique())

            # Apply color based on cost difference and new rules
            for i in range(len(sorted_costs) - 1):
                current_cost = sorted_costs[i]
                next_cost = sorted_costs[i + 1]
                gap = next_cost - current_cost

                # Apply blue if gap < 1.5
                if gap < 1.5 and cost == next_cost:
                    return 'blue'

                # Apply red if gap > 1.5 for the maximum value in this group
                elif gap > 1.5 and cost == sorted_costs[-1]:  # Highest cost in sorted list
                    return 'red'

            # Apply specific colors based on fixed costs, with precedence for high-to-low logic
            if cost == cost_min:
                return 'green'
            elif cost == cost_max:
                return 'red'
            elif cost <= cost_mean:
                return 'green'
            elif cost > cost_mean and qty <= cost_75th:
                return 'orange'

            # Default to gold for other cases
            return 'orange'

        # Apply the color code function to the DataFrame
        df_cleaned['Color Code'] = df_cleaned.apply(determine_color_code, axis=1)

        # Apply special condition for 'F0' rows to not have color
        condition = df_cleaned['Row Labels'].str.startswith(('F0', 'N0'))
        df_cleaned.loc[condition, 'Color Code'] = ' '  # Ensures that 'F0' rows have no specific color coding

        # Add the "Marking" column to assign 'X' for products with a red or orange sub-product
        df_cleaned['Marking'] = ''

        # Function to mark main and sub-products with 'X'
        def mark_products(group):
            if ('red') in group['Color Code'].values or ('orange') in group['Color Code'].values or ('blue') in group['Color Code'].values:
                group['Marking'] = 'X'
            return group

        # Apply marking to each product group
        df_cleaned = df_cleaned.groupby('Main Product').apply(mark_products)

        # Function to highlight cells based on the 'Color Code' column
        def highlight_color(row):
            if str(row['Row Labels']).startswith(('F0', 'N0')):
                return ['background-color: black; color:white; font-weight: bold'] * len(row)
            else:
                color = row['Color Code'] if 'Color Code' in row else 'green'  # Default to green if no code found
                return ['background-color: {}'.format(color)] * len(row)

        # Apply the color highlight to the DataFrame using Styler.apply
        styled_df = df_cleaned.style.apply(highlight_color, axis=1)

        # Save the styled DataFrame with adjusted column width
        save_styled_df_to_excel_with_width('styled_output.xlsx', styled_df)

        # Print the top 20 rows for visibility
        print("Top 20 rows of the new DataFrame with Color Code and Marking:")
        print(df_cleaned.head(20))

    else:
        result_label.configure(text="Please upload a file first.", text_color="red")

def download_file():
    if uploaded_file:
        # Get the file path for saving the Excel file
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", 
            filetypes=[("Excel files", "*.xlsx")],
            title="Save the highlighted Excel file"
        )
        
        # If the user has selected a file path
        if file_path:
            # Ensure the highlight_cells function has been called
            if 'styled_df' in globals():
                # Save the styled DataFrame as an Excel file
                styled_df.to_excel(file_path, index=False, engine='openpyxl')

                # Optionally, reopen the saved Excel file to modify additional formatting (e.g., bold for the Main Product column)
                wb = load_workbook(file_path)
                ws = wb.active

                # Example: Make the 'Main Product' column bold
                for idx in range(2, len(df_cleaned) + 2):  # Assuming df_cleaned is available
                    cell = ws[f'D{idx}']  # Assuming column D is the Main Product column
                    if cell.value:
                        cell.font = Font(bold=True)

                # Save the workbook again to retain the formatting
                wb.save(file_path)

                # Display success message in the result box
                result_box.insert(ctk.END, f"File saved to: {file_path}\n")
            else:
                result_box.insert(ctk.END, "Error: Highlighted data not found. Please run the process first.\n")
        else:
            result_box.insert(ctk.END, "Save operation was cancelled.\n")
    else:
        result_label.configure(text="Please upload a file first.", text_color="red")
        result_box.insert(ctk.END, "Please upload a file first.\n")


def reset():
    file_label.configure(text="No file uploaded", text_color=theme_color)
    result_label.configure(text="")
    global uploaded_file
    uploaded_file = None
    result_box.insert(ctk.END, "Reset completed.\n")

def open_settings():
    settings_frame.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")
    main_frame.grid_forget()

def close_settings():
    settings_frame.grid_forget()
    main_frame.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")

def switch_mode():
    if ctk.get_appearance_mode() == "Light":
        ctk.set_appearance_mode("Light")
        result_label.configure(text_color="white")
        result_box.configure(text_color="white")
    else:
        ctk.set_appearance_mode("Dark")
        result_label.configure(text_color="black")
        result_box.configure(text_color="black")
    save_settings()

def change_theme_color(color):
    global theme_color
    if color == "transparent":
        theme_color = "purple"  # Default color
    else:
        theme_color = color
    file_label.configure(text_color=theme_color)
    result_label.configure(text_color=theme_color)
    settings_button.configure(fg_color=theme_color)
    reset_button.configure(fg_color=theme_color)
    back_button.configure(fg_color=theme_color)
    switch_button.configure(fg_color=theme_color)
    edit_about_button.configure(fg_color=theme_color)
    for widget in main_frame.winfo_children():
        if isinstance(widget, ctk.CTkButton):
            widget.configure(fg_color=theme_color)
    bottom_line.configure(fg_color=theme_color)
    settings_bottom_line.configure(fg_color=theme_color)
    save_settings()

#Function to display About information


def load_about_text():
    """Load the About text from about.txt. Creates the file with default content if it doesn't exist."""
    if not os.path.exists('about.txt'):
        # If the file does not exist, create it with default content
        with open('about.txt', 'w') as file:
            default_content = ("Pdp Processor v1.0\n"
                               "Developed by Jiren Pandya\n"
                               "Other Contributor:\n"
                               "Applied by:\n"
                               "This software is the property of.\n"
                               "This software helps in analyzing Excel sheets.\n"
                               "In upcoming updates, we will provide time graph analysis, "
                               "bug fixes, and additional features.\n"
                               "The user ID is 0 and the password is 0.\n"
                               "The password and user ID are unchangeable.\n"
                               "For any help regarding this software, please call +91 8401751355 "
                               "(Jiren Pandya) or email at pandyajiren15@gmail.com.")
            file.write(default_content)

    # After ensuring the file exists, read its contents
    try:
        with open('about.txt', 'r') as file:
            return file.read().strip()
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while loading About text: {str(e)}")
        return ""

def show_about():
    """Display the About section from about.txt."""
    try:
        load_about_text()
        # Load About text using the load_about_text function
        about_text = load_about_text()
        
        # Show the About text in a message box
        messagebox.showinfo("About", about_text)

    except Exception as e:
        # Handle any other exceptions
        messagebox.showerror("Error", f"An error occurred: {str(e)}")




user_db = {}
is_signed_in=False
# File to store user credentials
USER_DB_FILE = "user_credentials.txt"

def create_default_users():
    """Create default users if the user database is empty."""
    default_users = {
        "admin": "admin@123",
        "developer": "dev@123",
        "0":"0"
    }
    with open(USER_DB_FILE, "w") as file:
        for user_id, password in default_users.items():
            hashed_password = hash_password(password)
            file.write(f"{user_id},{hashed_password}\n")

def load_user_db():
    """Load user data from the text file into the user_db dictionary."""

    if not os.path.exists(USER_DB_FILE):
        create_default_users()

    if os.path.getsize(USER_DB_FILE) == 0:
        create_default_users()  # Create default users if the file is empty
        
    try:
        with open(USER_DB_FILE, "r") as file:
            for line in file:
                user_id, hashed_password = line.strip().split(',')
                user_db[user_id] = hashed_password
    except FileNotFoundError:
        # File does not exist yet, no users have been registered
        pass

def save_user_db(user_id, hashed_password):
    """Append the new user's credentials to the text file."""
    with open(USER_DB_FILE, "a") as file:
        file.write(f"{user_id},{hashed_password}\n")

def hash_password(password):
    """Hash the password using SHA-256 for secure storage."""
    return hashlib.sha256(password.encode()).hexdigest()


def update_admin_controls():
    if is_admin or is_dev:
        edit_about_button.grid()  # Show the button if is_admin is True
    else:
        edit_about_button.grid_remove()  # Hide the button if not admin
   
is_admin = False
is_dev=False     
def authenticate_user():
    def check_credentials():
        global is_admin,is_dev
        user_id = user_id_entry.get()
        password = hash_password(password_entry.get())  # Hash the entered password
        
        if user_id in user_db and user_db[user_id] == password:
            
            if user_id=="admin":
                is_admin=True
                update_admin_controls()
            if user_id=="developer":
                is_dev=True
                print("admin/developer signed in")
                update_admin_controls()


            auth_window.destroy()
            app.deiconify()  # Show the main app window
        else:
            messagebox.showerror("Error", "Invalid credentials")


    def register_user():
        user_id = user_id_entry.get()
        password = password_entry.get()
        if user_id in user_db:
            messagebox.showerror("Error", "User ID already exists. Please login.")
        elif not user_id or not password:
            messagebox.showerror("Error", "Please fill in both fields.")
        else:
            hashed_password = hash_password(password)
            user_db[user_id] = hashed_password  # Store the hashed password in the dictionary
            save_user_db(user_id, hashed_password)  # Save the new user to the file
            messagebox.showinfo("Success", "Registration successful! Please log in.")
            clear_entries()  # Clear the input fields
            show_login_window()  # After registration, return to login

    def show_register_window():
        """Switch the window to show registration fields."""
        clear_entries()
        login_button.pack_forget()
        ok_button.pack(pady=10)
        register_button.configure(text="Register", command=register_user)
        switch_to_login_button.pack(side=tk.BOTTOM)

    def show_login_window():
        """Switch back to the login view."""
        clear_entries()
        ok_button.pack_forget()
        register_button.pack_forget()
        login_button.pack(pady=10)
        switch_to_register_button.pack(side=tk.BOTTOM)

    def clear_entries():
        """Clear both the User ID and Password fields."""
        user_id_entry.delete(0, tk.END)
        password_entry.delete(0, tk.END)

    # Load user data from the file
    load_user_db()

    # Create authentication window
    auth_window = Toplevel(app)
    auth_window.title("Authentication")
    auth_window.geometry("900x350")
    auth_window.resizable(False, False)

    # Create User ID and Password fields
    ctk.CTkLabel(auth_window, text="User ID:", text_color="black").pack(pady=5)
    user_id_entry = ctk.CTkEntry(auth_window,)
    user_id_entry.pack(pady=5)

    ctk.CTkLabel(auth_window, text="Password:",text_color="black").pack(pady=5)
    password_entry = ctk.CTkEntry(auth_window, show="*")
    password_entry.pack(pady=5)

    # Create Login button
    login_button = ctk.CTkButton(auth_window, text="Login", command=check_credentials)

    # Create "OK" button for registration
    ok_button = ctk.CTkButton(auth_window, text="OK", command=register_user)

    # Create Register button
    register_button = ctk.CTkButton(auth_window, text="Register", command=register_user)

    # Create switch buttons to toggle between Login and Register
    switch_to_register_button = ctk.CTkButton(auth_window, text="Don't have an account? Register here", command=show_register_window)
    switch_to_register_button.pack(side=tk.BOTTOM)

    switch_to_login_button = ctk.CTkButton(auth_window, text="Already have an account? Login here", command=show_login_window)

    # Start with the Login view
    login_button.pack(pady=10)




def save_about_text(about_text):
    """Save the About text to about.txt."""
    with open('about.txt', 'w') as file:
        file.write(about_text)

def edit_about():
    global about_text
    global is_admin, is_dev  # Ensure these are defined and initialized elsewhere

    if is_admin or is_dev:
        edit_window = Toplevel()
        edit_window.title("Edit About")

        # Create a Text widget to edit the About text
        text_editor = Text(edit_window, wrap="word")
        text_editor.insert("1.0", about_text)  # Insert the current About text
        text_editor.pack(expand=True, fill="both")

        # Store the original text to revert back if needed
        original_text = about_text

    # Save changes made by admin
    def save_changes():
        global about_text
        try:
            current_text = text_editor.get("1.0", "end").strip()
            about_text = text_editor.get("1.0", "end").strip()

            # Define required phrases for case-insensitive checking
            required_strings = [
                "developed by jiren pandya",
                "for any help regarding this software, please call +91 8401751355 (jiren pandya) or email at pandyajiren15@gmail.com"
            ]

                           # Check if the current text is empty or missing required phrases
            if not current_text:  # If the text editor is empty
                    messagebox.showwarning("Warning", "The About section cannot be empty.")
                    text_editor.delete("1.0", "end")  # Clear the editor
                    text_editor.insert("1.0", original_text)  # Reset to original text
                    return

            if not is_dev: #only developer can remove jiren name 
                # Verify that required phrases are still present (case-insensitive)
                if not all(req_str in about_text.lower() for req_str in required_strings):
                    messagebox.showwarning("Warning", "The name 'Jiren Pandya' and contact information cannot be removed or modified.")
                    text_editor.delete("1.0", "end")  # Clear the current text in the editor
                    text_editor.insert("1.0", original_text)  # Reinsert the original text
                    return  # Prevent saving if any required phrase is missing

            # Save changes to about.txt if checks pass
            save_about_text(about_text)
            messagebox.showinfo("Success", "About section updated successfully.")
            edit_window.destroy()
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while saving: {str(e)}")

    # Add a button to save changes
    save_button = ctk.CTkButton(edit_window, text="Save", command=save_changes)
    save_button.pack(pady=10)



app = ctk.CTk()
app.title("Paint Details Processor")
app.geometry("955x645")

# Disable resizing the window
app.resizable(False, False)
app.withdraw()  # Hide the main app window initially

uploaded_file = None
theme_color = "purple"

# Load settings
load_settings()

# Authenticate user
authenticate_user()

# Main Frame
main_frame = ctk.CTkFrame(app)
main_frame.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")

# Upload button
upload_button = ctk.CTkButton(main_frame, text="Upload Excel Sheet", command=upload_file, fg_color=theme_color, width=130, height=50)
upload_button.grid(row=0, column=0, padx=10, pady=10)

# # Process button
# process_button = ctk.CTkButton(main_frame, text="   Process File  ", command=process_file, fg_color=theme_color, width=130, height=50)
# process_button.grid(row=1, column=0, padx=10, pady=10)

# Highlight button
highlight_button = ctk.CTkButton(main_frame, text="Highlight Cells", command=highlight_cells, fg_color=theme_color, width=130, height=50)
highlight_button.grid(row=1, column=0, padx=10, pady=10)

# Download button
download_button = ctk.CTkButton(main_frame, text="Download File", command=download_file, fg_color=theme_color, width=130, height=50)
download_button.grid(row=2, column=0, padx=10, pady=10)

# File label
file_label = ctk.CTkLabel(main_frame, text="No file uploaded", text_color=theme_color)
file_label.grid(row=3, column=0, padx=10, pady=10)

# Result label
result_label = ctk.CTkLabel(main_frame, text="", text_color="black")
result_label.grid(row=4, column=0, padx=10, pady=10)

# Result# Result box
border_frame = ctk.CTkFrame(main_frame, corner_radius=10, border_width=2, border_color="grey")
border_frame.grid(row=0, column=1, rowspan=6, padx=10, pady=10)

# Create a Textbox inside the frame
result_box = ctk.CTkTextbox(border_frame, height=500, width=580)
result_box.pack(fill="both", expand=True, padx=5, pady=5) 
# Settings button with icon
settings_button = ctk.CTkButton(main_frame, text="⚙", command=open_settings, fg_color=theme_color, width=50, height=50)
settings_button.grid(row=0, column=2, padx=10, pady=10)

# Reset button with icon
reset_button = ctk.CTkButton(main_frame, text="⟳", command=reset, fg_color=theme_color, width=50, height=50)
reset_button.grid(row=0, column=3, padx=10, pady=10)

# Bottom line
bottom_line = ctk.CTkFrame(main_frame, height=2, fg_color=theme_color)
bottom_line.grid(row=6, column=0, columnspan=4, padx=10, pady=5, sticky="ew")

# Logo

# Load and resize the image
#logo_image = Image.open("C:/Users/SMIT/OneDrive/Desktop/jiren/PP/pdp.png")
#logo_image = logo_image.resize((100, 100), Image.LANCZOS)  # Adjust size as needed

# Ensure that the image has an alpha channel for transparency
#logo_image = logo_image.convert("RGBA")

# Convert to a format suitable for tkinter
#logo_photo = ImageTk.PhotoImage(logo_image)

# Create a label and set the image
#logo_label = tk.Label(main_frame, image=logo_photo, bg="white")  # Use any solid background if needed
#logo_label.image = logo_photo  # Keep a reference to avoid garbage collection
#logo_label.grid(row=6, column=3, padx=10, pady=5, sticky="e")
# Settings Frame
settings_frame = ctk.CTkFrame(app)

# Back button
back_button = ctk.CTkButton(settings_frame, text="<", command=close_settings, fg_color=theme_color,width=50, height=50)
back_button.grid(row=0, column=0, padx=10, pady=10, sticky="nw")

# Switch mode button
switch_button = ctk.CTkSwitch(settings_frame, text="Dark/Light Mode", command=switch_mode)
switch_button.grid(row=1, column=0, padx=10, pady=10, sticky="nw")



# Theme color options
theme_label = ctk.CTkLabel(settings_frame, text="Select Theme Color")
theme_label.grid(row=2, column=0, padx=10, pady=10, sticky="nw")

colors = [
    "black", "darkblue", "darkred", "dark green", "darkslategrey", "SlateGray",
    "Indigo", "DarkViolet", "SaddleBrown", "olive", "tan", "blue", "dodger blue",
    "#007BFF", "cyan4", "green", "lime green", "orange", "red", "purple",
    "deep pink", "transparent"
     ]
color_buttons = ctk.CTkFrame(settings_frame)
color_buttons.grid(row=3, column=0, padx=10, pady=10, sticky="nw")

for color in colors:
    color_button = ctk.CTkButton(color_buttons, text="✔" if color == theme_color else "", fg_color=color if color != "transparent" else "white", width=30, height=30, command=lambda c=color: change_theme_color(c))
    color_button.pack(side="left", padx=5)

# About button
about_button = ctk.CTkButton(settings_frame, text="About", command=show_about, fg_color=theme_color,width=130, height=50)
about_button.grid(row=4, column=0, padx=10, pady=10, sticky="nw")

# Edit About button (only visible if admin or dev is logged in)
edit_about_button = ctk.CTkButton(settings_frame, text="Edit About", command=edit_about, fg_color=theme_color, width=130, height=50)

if is_admin or is_dev:
    edit_about_button.grid(row=4, column=1, padx=10, pady=10, sticky="nw")  # Show the button if admin or dev


default_cost_column = "B"
default_qty_column = "C"
default_sheet_name = "Sheet2"

# Default Cost column for processing
cost_column_var = StringVar(value=default_cost_column)
cost_column_label = ctk.CTkLabel(settings_frame, text="Default Cost Column ")
cost_column_label.grid(row=5, column=0, padx=10, pady=10, sticky="nw")
cost_column_entry = ctk.CTkEntry(settings_frame, textvariable=cost_column_var)
cost_column_entry.grid(row=6, column=0, padx=10, pady=10, sticky="nw")

# Default Quantity column for processing
qty_column_var = StringVar(value=default_qty_column)
qty_column_label = ctk.CTkLabel(settings_frame, text="Default Quantity Column ")
qty_column_label.grid(row=7, column=0, padx=10, pady=10, sticky="nw")
qty_column_entry = ctk.CTkEntry(settings_frame, textvariable=qty_column_var)
qty_column_entry.grid(row=8, column=0, padx=10, pady=10, sticky="nw")

# Default Sheet name for processing
Sheet_var = StringVar(value=default_sheet_name)
Sheet_name='Sheet2'
Sheet_label = ctk.CTkLabel(settings_frame, text="Default Sheet for Processing")
Sheet_label.grid(row=9, column=0, padx=10, pady=10, sticky="nw")
Sheet_entry = ctk.CTkEntry(settings_frame, textvariable=Sheet_var)
Sheet_entry.grid(row=10, column=0, padx=10, pady=10, sticky="nw")


#pimage = Image.open("C:/Users/SMIT/OneDrive/Desktop/jiren/PP/b.png")
#pimage = pimage.resize((100, 100), Image.LANCZOS)  # Adjust size as needed

# Ensure that the image has an alpha channel for transparency
#pimage = pimage.convert("RGBA")

# Convert to a format suitable for tkinter
#pphoto = ImageTk.PhotoImage(pimage)

# Create a label and set the image
#plabel = tk.Label(settings_frame, image=pphoto, bg="white")  # Use any solid background if needed
#plabel.image = pphoto  # Keep a reference to avoid garbage collection
#plabel.grid(row=6, column=0, padx=10, pady=5, sticky="e")
# Bottom line for settings frame
settings_bottom_line = ctk.CTkFrame(settings_frame, height=2, fg_color=theme_color)
settings_bottom_line.grid(row=11, column=0, columnspan=2, padx=10, pady=5, sticky="ew")

app.mainloop()